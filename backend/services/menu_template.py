"""메뉴 일괄 등록용 Excel 양식 생성 서비스.

GET /api/tools/menu-template.xlsx (Basic+)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 열 정의: (헤더, 너비, 예시값)
_COLUMNS = [
    ("메뉴명 *", 22, "아메리카노"),
    ("가격 (원) *", 14, "4500"),
    ("설명", 40, "신선한 원두로 추출한 에스프레소에 물을 더한 기본 커피"),
    ("카테고리", 18, "음료"),
    ("대표메뉴 (Y/N)", 16, "Y"),
    ("알레르기 정보", 22, "우유(선택)"),
    ("칼로리 (kcal)", 16, "10"),
    ("비고", 30, "시즌 한정 메뉴"),
]

_SAMPLE_ROWS = [
    ["아메리카노", "4500", "신선한 원두로 추출한 기본 커피", "음료", "Y", "", "10", ""],
    ["카페라떼", "5000", "에스프레소 + 스팀 우유", "음료", "Y", "우유", "120", ""],
    ["치즈케이크", "6500", "부드러운 뉴욕 스타일 치즈케이크", "디저트", "N", "우유,달걀,밀", "320", "매일 소량 제조"],
    ["아보카도 토스트", "8000", "통밀 브레드에 신선한 아보카도", "푸드", "Y", "밀,달걀", "410", ""],
]

_HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
_HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
_REQUIRED_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
_REQUIRED_FONT = Font(name="맑은 고딕", bold=True, color="1E40AF", size=10)
_SAMPLE_FONT = Font(name="맑은 고딕", color="374151", size=10)
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_menu_template() -> bytes:
    """메뉴 양식 Excel 파일을 생성하여 bytes로 반환."""
    wb = Workbook()
    ws = wb.active
    ws.title = "메뉴 목록"

    # 안내 행 (1행)
    ws.merge_cells("A1:H1")
    ws["A1"] = "★ * 표시 항목은 필수입니다. 작성 후 담당자에게 전달해주세요."
    ws["A1"].font = Font(name="맑은 고딕", bold=True, color="DC2626", size=11)
    ws["A1"].fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 22

    # 헤더 행 (2행)
    for col_idx, (header, width, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 22

    # 예시 행 (3행) — 회색 배경으로 예시임을 표시
    example_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    example_font = Font(name="맑은 고딕", color="6B7280", italic=True, size=10)
    ws.merge_cells("A3:H3")
    ws["A3"] = "▼ 아래는 예시입니다. 삭제 후 실제 메뉴를 입력해주세요."
    ws["A3"].font = Font(name="맑은 고딕", italic=True, color="6B7280", size=10)
    ws["A3"].fill = example_fill
    ws["A3"].alignment = _CENTER
    ws.row_dimensions[3].height = 18

    # 샘플 데이터 (4~7행)
    for row_idx, row_data in enumerate(_SAMPLE_ROWS, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.fill = example_fill
            cell.alignment = _LEFT if col_idx in (3, 8) else _CENTER
            cell.border = _BORDER
        ws.row_dimensions[row_idx].height = 20

    # 입력 영역 안내 행 (8행)
    ws.merge_cells("A8:H8")
    ws["A8"] = "▼ 이 아래 행에 실제 메뉴를 입력해주세요."
    ws["A8"].font = Font(name="맑은 고딕", bold=True, color="1D4ED8", size=10)
    ws["A8"].fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    ws["A8"].alignment = _CENTER
    ws.row_dimensions[8].height = 18

    # 입력 행 9~58 (50개 빈 행)
    for row_idx in range(9, 59):
        for col_idx in range(1, len(_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # 필수 열은 연한 파랑 배경
            if col_idx in (1, 2):
                cell.fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
            cell.border = _BORDER
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = _LEFT if col_idx in (3, 8) else _CENTER
        ws.row_dimensions[row_idx].height = 20

    # 틀 고정 (3행 이하 스크롤)
    ws.freeze_panes = "A9"

    # 시트 설정
    ws.sheet_view.showGridLines = True
    ws.print_area = f"A1:H{58}"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
